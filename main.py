from openpyxl import load_workbook
from latex_content import LatexContent as lc
from latexcompiler import LC as compiler

def extract_excel_data(file_path):
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)

    
    # Access the specific sheet
    sheet = wb['Dokumentausgangliste']
    
    # Initialize list to store results
    results = []
    
    # Define column indices (0-based)
    col_ag_idx = 32  # Column AG
    col_n_idx = 13   # Column N
    col_al_idx = 37   # Column AL
    col_ao_idx = 40   # Column AO
    col_ar_idx = 43  # Column AR
    col_as_idx = 44  # Column AS
    
    row_count = 0
    processed_count = 0

    for row in sheet.iter_rows(min_row=17):
        row_count += 1
        
        try:
            # Get AG cell value first
            ag_value = row[col_ag_idx].value
            
            
            # If AG column has a value, collect the data
            if ag_value:
                processed_count += 1
                n_value = row[col_n_idx].value
                al_value = row[col_al_idx].value
                ao_value = row[col_ao_idx].value
                ar_value = row[col_ar_idx].value
                as_value = row[col_as_idx].value

                
                results.append({
                    "proj_name": n_value,
                    "page_type": al_value,
                    "proj_code": ao_value,
                    "author": ar_value,
                    "checker": as_value,
                    "rev_num": ag_value
                })
                
        except IndexError as e:
            print(f"Row {row_count + 16} doesn't have enough columns. Available columns: {len(row)}")
            continue
        except Exception as e:
            print(f"Error processing row {row_count + 16}: {str(e)}")
            continue

    wb.close()
    
    return results

def create_report(output_folder, output_file,report_name,proj_code,rev_num,author, description, checker,auth_init, check_init, author_email,page_type):
    with open(output_file, "a", encoding="utf-8") as tex_file:
       tex_file.write(lc.content(r"C:/Users/vmylavarapu/Desktop/ProjectTestFolder/BH.jpg".replace('/','\\'),"Museum für Naturkunde Berlin",report_name,proj_code, "0057074" ,rev_num, author, description, checker, auth_init, check_init, author_email,page_type))
    compiler.compile_document(tex_engine='lualatex',
                            bib_engine='biber',
                            no_bib=True,
                            path=output_file, folder_name= output_folder)

    print("Done!")   


    return None

# Example usage:
if __name__ == "__main__":
    try:
        file_path = r"C:\Users\vmylavarapu\Downloads\BHE_DE_Dokumentenregister_MfN_LP2.xlsx"

        data = extract_excel_data(file_path)
        
        for item in data:
            auth_name = item['author']
            check_name = item['checker']
            words = auth_name.split()
            auth_init = ''.join([word[0] for word in words]) 
            words = check_name.split()
            check_init = ''.join([word[0] for word in words]) 
            email_id = auth_name.lower().replace(' ', '.') + "@burohappold.com"
            
            # Create a sensible file name
            file_name = f"{item['proj_name'].replace(' ', '_')}_{item['proj_code']}.tex"
            folder_name = f"{item['proj_name'].replace(' ', '_')}_{item['proj_code']}"
            
            create_report(folder_name, file_name, item['proj_name'], item['proj_code'].replace('_','\\_'), item['rev_num'].replace('_','\\_'), auth_name, item['proj_name'], check_name, auth_init, check_init, email_id, item['page_type'])
        
        if not data:
            print("No data was extracted! Please check the debug output above.")
            
    except Exception as e:
        print(f"An error occurred: {str(e)}")

